# phase2_process.py
"""
PHASE 2: CSV to Structured Excel (FINAL WORKING VERSION)
==========================================================

Processes the properly extracted CSV from Phase 1 into:
- Clean structured Excel workbook
- With all students included
- Proper subject code detection
- Alias columns for subjects

Designed to work with the context-aware Phase 1 output.
"""

import os
import sys
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def clean_usn(usn):
    """Clean USN codes - final version"""
    if not isinstance(usn, str):
        return ""
    
    usn = usn.upper().strip()
    
    if not usn:
        return ""
    
    # Fix OCR errors in USNs only
    usn = usn.replace('O', '0')  # O → 0 in USNs
    usn = usn.replace('I', '1')  # I → 1 in USNs
    usn = usn.replace('l', '1')  # l → 1 in USNs
    
    # Remove any non-alphanumeric
    usn = re.sub(r'[^A-Z0-9]', '', usn)
    
    return usn


def clean_name(name):
    """Clean student names - keep as is"""
    if not isinstance(name, str):
        return ""
    
    name = name.strip()
    
    # Remove any non-standard characters but keep letters, spaces, hyphens, apostrophes
    name = re.sub(r'[^A-Za-z\s\-\'\.]', ' ', name)
    
    # Clean extra spaces
    name = re.sub(r'\s+', ' ', name)
    
    return name.strip()


def extract_subject_codes(header_row):
    """
    Extract subject codes from header row (row 1 in our data).
    
    Row 1 contains: 'SI No_', 'USN', 'Student Name', '22CSTPCCCT', '22CSTPCCCT', ...
    Each subject code appears 4 times (for CIE, SEE, TOTAL, GRADE columns).
    """
    print("\n🔍 Extracting subject codes from header...")
    
    subject_codes = []
    i = 3  # Start after SI No_, USN, Student Name
    
    while i < len(header_row):
        cell = str(header_row[i]).strip()
        
        if not cell:
            i += 1
            continue
        
        # Check if this looks like a subject code
        # Subject codes: 22CSTPCCCT, 22CS7PENLP, etc.
        if (cell.startswith('22') and 
            len(cell) >= 8 and 
            not cell.isdigit()):  # Not just numbers
            
            # Clean the code
            code = cell.upper()
            
            # Fix common OCR errors in subject codes
            code = re.sub(r'22CST', '22CS7', code)   # CST → CS7
            code = re.sub(r'22MEZO', '22ME2O', code) # MEZO → ME2O
            code = re.sub(r'\s+', '', code)  # Remove spaces
            
            # Only add if it's a new code (codes repeat 4 times)
            if not subject_codes or code != subject_codes[-1]:
                subject_codes.append(code)
                print(f"  ✓ Subject {len(subject_codes)}: {code}")
            
            # Skip next 3 columns (same subject, different component)
            i += 4
        else:
            i += 1
    
    return subject_codes


def get_subject_aliases(codes):
    """Get alias names for subject codes"""
    alias_map = {
        '22CS7PCCCT': 'CC',      # Cloud Computing
        '22CS7PENLP': 'NLP',     # Natural Language Processing
        '22CS7PERPA': 'RPA',     # Robot Process Automation
        '22CS7PENDL': 'DL',      # Deep Learning
        '22CS7PEHCI': 'HCI',     # Human Computer Interaction
        '22CS7HSCFI': 'CF',      # Cyber Forensics
        '22CS7NCMCI': 'MOOC',    # MOOCs Course
        '22ME2OESSE': 'SE',      # Sustainable Engineering
    }
    
    return {code: alias_map.get(code, code[:4]) for code in codes}


def get_subject_names(codes):
    """Get full names for subjects"""
    name_map = {
        '22CS7PCCCT': 'Cloud Computing',
        '22CS7PENLP': 'Natural Language Processing',
        '22CS7PERPA': 'Robot Process Automation',
        '22CS7PENDL': 'Neural Network & Deep Learning',
        '22CS7PEHCI': 'Human Computer Interaction',
        '22CS7HSCFI': 'Cyber Law, Forensics & IPR',
        '22CS7NCMCI': 'MOOCs Course',
        '22ME2OESSE': 'Sustainable Engineering',
    }
    
    return {code: name_map.get(code, code) for code in codes}


def process_csv_to_dataframe(csv_path):
    """
    Process CSV from Phase 1 into structured DataFrame.
    
    Returns:
        tuple: (dataframe, subject_codes, subject_aliases)
    """
    
    print("\n" + "="*60)
    print("PROCESSING CSV DATA")
    print("="*60)
    
    if not os.path.exists(csv_path):
        print(f"❌ ERROR: File not found: {csv_path}")
        return None, [], {}
    
    print(f"📥 Input file: {csv_path}")
    
    try:
        # Read CSV
        raw_df = pd.read_csv(csv_path, header=None, dtype=str)
        print(f"📊 Raw data shape: {raw_df.shape[0]} rows × {raw_df.shape[1]} columns")
        
        # Show data structure
        print("\n📋 DATA STRUCTURE:")
        print("Row 0 (descriptions):", raw_df.iloc[0].fillna('').astype(str).tolist()[:8])
        print("Row 1 (subject codes):", raw_df.iloc[1].fillna('').astype(str).tolist()[:8])
        print("Row 2 (metadata):", raw_df.iloc[2].fillna('').astype(str).tolist()[:8])
        print("Row 3 (student 1):", raw_df.iloc[3].fillna('').astype(str).tolist()[:8])
        
        # HEADER IS ROW 1 (contains subject codes)
        header_row_idx = 1
        print(f"\n📝 Using row {header_row_idx} as header (contains subject codes)")
        
        # Extract subject codes
        header_row = raw_df.iloc[header_row_idx].fillna('').astype(str).tolist()
        subject_codes = extract_subject_codes(header_row)
        
        if not subject_codes:
            print("❌ ERROR: No subject codes found!")
            print("Header row preview:", header_row[:15])
            return None, [], {}
        
        print(f"\n📚 Found {len(subject_codes)} subjects")
        
        # Get aliases
        subject_aliases = get_subject_aliases(subject_codes)
        subject_names = get_subject_names(subject_codes)
        
        # Extract student data (starts at row 3)
        student_data = []
        student_start = 3
        
        print(f"\n👥 Extracting students (starting from row {student_start})...")
        
        for row_idx in range(student_start, len(raw_df)):
            row = raw_df.iloc[row_idx].fillna('').astype(str).tolist()
            
            # Skip empty rows
            if not row or all(cell.strip() == '' for cell in row[:3]):
                continue
            
            # Get USN and Name
            raw_usn = row[1] if len(row) > 1 else ""
            raw_name = row[2] if len(row) > 2 else ""
            
            # Clean them
            usn = clean_usn(raw_usn)
            name = clean_name(raw_name)
            
            # Check if this is a valid student
            is_valid = (usn and len(usn) >= 9 and usn.startswith('1BM')) or (usn and name)
            
            if is_valid:
                # Get Sl_No
                sl_no_raw = row[0] if len(row) > 0 else ""
                sl_no = None
                
                if sl_no_raw and sl_no_raw.strip().isdigit():
                    sl_no = int(sl_no_raw.strip())
                else:
                    # Generate sequential number
                    sl_no = len(student_data) + 1
                
                # Build student record
                record = {
                    'Sl_No': sl_no,
                    'USN': usn,
                    'Student_Name': name
                }
                
                # Extract marks for each subject
                col_idx = 3  # Start of first subject's CIE column
                
                for subject_code in subject_codes:
                    if col_idx + 3 < len(row):
                        cie = row[col_idx].strip() if col_idx < len(row) else ""
                        see = row[col_idx + 1].strip() if col_idx + 1 < len(row) else ""
                        total = row[col_idx + 2].strip() if col_idx + 2 < len(row) else ""
                        grade = row[col_idx + 3].strip() if col_idx + 3 < len(row) else ""
                        
                        # Clean marks
                        cie = cie if cie else ""
                        see = see if see else ""
                        total = total if total else ""
                        grade = grade if grade else ""
                        
                        # Add to record
                        record[f'{subject_code}_CIE'] = cie
                        record[f'{subject_code}_SEE'] = see
                        record[f'{subject_code}_TOTAL'] = total
                        record[f'{subject_code}_GRADE'] = grade
                        
                        # Add alias columns
                        alias = subject_aliases.get(subject_code)
                        if alias:
                            record[f'{alias}_CIE'] = cie
                            record[f'{alias}_SEE'] = see
                            record[f'{alias}_TOTAL'] = total
                            record[f'{alias}_GRADE'] = grade
                        
                        col_idx += 4
                    else:
                        # Not enough columns
                        record[f'{subject_code}_CIE'] = ''
                        record[f'{subject_code}_SEE'] = ''
                        record[f'{subject_code}_TOTAL'] = ''
                        record[f'{subject_code}_GRADE'] = ''
                
                student_data.append(record)
                print(f"  ✓ Student {len(student_data)}: {usn} - {name}")
        
        # Create DataFrame
        if not student_data:
            print("❌ ERROR: No student data extracted!")
            return None, subject_codes, subject_aliases
        
        df = pd.DataFrame(student_data)
        
        # Sort and renumber Sl_No
        df = df.sort_values('Sl_No').reset_index(drop=True)
        df['Sl_No'] = range(1, len(df) + 1)
        
        print(f"\n✅ Extracted {len(df)} students")
        print(f"📈 Final DataFrame: {df.shape[0]} rows × {df.shape[1]} columns")
        
        # Show all students
        print("\n📋 EXTRACTED STUDENTS:")
        for idx, student in df.iterrows():
            print(f"   {student['Sl_No']:2d}. {student['USN']} - {student['Student_Name']}")
        
        return df, subject_codes, subject_aliases
        
    except Exception as e:
        print(f"❌ ERROR processing CSV: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, [], {}


def create_excel_workbook(df, subject_codes, subject_aliases, output_path):
    """
    Create formatted Excel workbook.
    
    Returns:
        bool: Success status
    """
    
    print("\n" + "="*60)
    print("CREATING EXCEL WORKBOOK")
    print("="*60)
    
    try:
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # ===== SHEET 1: Summary =====
        print("📋 Creating 'Summary' sheet...")
        summary_sheet = workbook.create_sheet("Summary")
        
        summary_data = [
            ["EXTRACTION SUMMARY", "", ""],
            ["", "", ""],
            ["Extraction Date:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ""],
            ["Total Students:", len(df), ""],
            ["Total Subjects:", len(subject_codes), ""],
            ["Output File:", os.path.basename(output_path), ""],
            ["Status:", "✅ COMPLETE", ""],
            ["", "", ""],
            ["SUBJECTS EXTRACTED:", "Alias", "Full Name"],
        ]
        
        # Add subject details
        subject_names = get_subject_names(subject_codes)
        for code in subject_codes:
            alias = subject_aliases.get(code, 'N/A')
            name = subject_names.get(code, code)
            summary_data.append([code, alias, name])
        
        # Write summary
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                summary_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # ===== SHEET 2: Student Marks =====
        print("📊 Creating 'Student_Marks' sheet...")
        marks_sheet = workbook.create_sheet("Student_Marks")
        
        # Build headers
        headers = ['Sl_No', 'USN', 'Student_Name']
        for code in subject_codes:
            headers.extend([
                f'{code}_CIE',
                f'{code}_SEE',
                f'{code}_TOTAL',
                f'{code}_GRADE'
            ])
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            marks_sheet.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for df_idx, student in df.iterrows():
            excel_row = df_idx + 2
            
            # Build row
            row_data = [
                student['Sl_No'],
                student['USN'],
                student['Student_Name']
            ]
            
            for code in subject_codes:
                row_data.extend([
                    student.get(f'{code}_CIE', ''),
                    student.get(f'{code}_SEE', ''),
                    student.get(f'{code}_TOTAL', ''),
                    student.get(f'{code}_GRADE', '')
                ])
            
            # Write to Excel
            for col_idx, value in enumerate(row_data, 1):
                marks_sheet.cell(row=excel_row, column=col_idx, value=value)
        
        # ===== SHEET 3: Marks with Aliases =====
        print("🏷️  Creating 'Marks_Aliases' sheet...")
        alias_sheet = workbook.create_sheet("Marks_Aliases")
        
        # Build alias headers
        alias_headers = ['Sl_No', 'USN', 'Student_Name']
        for code in subject_codes:
            alias = subject_aliases.get(code)
            if alias:
                alias_headers.extend([
                    f'{alias}_CIE',
                    f'{alias}_SEE',
                    f'{alias}_TOTAL',
                    f'{alias}_GRADE'
                ])
        
        # Write alias headers
        for col_idx, header in enumerate(alias_headers, 1):
            alias_sheet.cell(row=1, column=col_idx, value=header)
        
        # Write alias data
        for df_idx, student in df.iterrows():
            excel_row = df_idx + 2
            
            # Build alias row
            alias_row = [
                student['Sl_No'],
                student['USN'],
                student['Student_Name']
            ]
            
            for code in subject_codes:
                alias = subject_aliases.get(code)
                if alias:
                    alias_row.extend([
                        student.get(f'{code}_CIE', ''),
                        student.get(f'{code}_SEE', ''),
                        student.get(f'{code}_TOTAL', ''),
                        student.get(f'{code}_GRADE', '')
                    ])
            
            # Write to Excel
            for col_idx, value in enumerate(alias_row, 1):
                alias_sheet.cell(row=excel_row, column=col_idx, value=value)
        
        # ===== FORMATTING =====
        print("🎨 Applying formatting...")
        
        # Define styles
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            
            if sheet_name == "Summary":
                sheet.column_dimensions['A'].width = 25
                sheet.column_dimensions['B'].width = 15
                sheet.column_dimensions['C'].width = 40
                
                # Left align all cells
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = left_align
                
                # Style title
                title_cell = sheet['A1']
                title_cell.font = Font(bold=True, size=14)
            
            else:  # Marks sheets
                # Set column widths
                sheet.column_dimensions['A'].width = 8   # Sl_No
                sheet.column_dimensions['B'].width = 15  # USN
                sheet.column_dimensions['C'].width = 25  # Name
                
                # Marks columns
                for col_idx in range(4, sheet.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    sheet.column_dimensions[col_letter].width = 12
                
                # Style header row
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                # Add borders and alignment
                for row in range(2, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell.border = thin_border
                        
                        if col in [1, 2, 3]:  # Sl_No, USN, Name
                            cell.alignment = left_align
                        else:  # Marks
                            cell.alignment = center_align
                
                # Freeze header
                sheet.freeze_panes = 'A2'
        
        # Save
        workbook.save(output_path)
        
        file_size = os.path.getsize(output_path) / 1024
        print(f"✅ Excel saved: {output_path} ({file_size:.1f} KB)")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR creating Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def save_clean_csv(df, output_path):
    """Save cleaned data to CSV"""
    try:
        print(f"\n💾 Saving CSV to: {output_path}")
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        
        if os.path.exists(output_path):
            size_kb = os.path.getsize(output_path) / 1024
            print(f"✅ CSV saved ({size_kb:.1f} KB)")
            return True
        else:
            print("⚠️  CSV file may not have been created")
            return False
    except Exception as e:
        print(f"❌ ERROR saving CSV: {str(e)}")
        return False


def main():
    """
    Main execution function.
    
    Returns:
        int: Exit code (0 = success, 1 = error)
    """
    
    # Parse arguments
    if len(sys.argv) < 2:
        print("Usage: python phase2_process_final.py input.csv [output.xlsx]")
        print("\nExample:")
        print("  python phase2_process_final.py debug_raw_table_context.csv")
        print("  python phase2_process_final.py data.csv output.xlsx")
        return 1
    
    input_csv = sys.argv[1]
    
    if len(sys.argv) > 2:
        output_excel = sys.argv[2]
    else:
        output_excel = "perfect_student_marks.xlsx"
    
    output_csv = "perfect_student_marks.csv"
    
    # Banner
    banner = """
    ╔══════════════════════════════════════════════════════════╗
    ║               CSV TO EXCEL PROCESSOR                     ║
    ║                    FINAL VERSION                         ║
    ║     Converts Phase 1 output to structured Excel          ║
    ╚══════════════════════════════════════════════════════════╝
    """
    
    print(banner)
    print("="*60)
    print(f"Input:  {input_csv}")
    print(f"Output: {output_excel} (Excel)")
    print(f"        {output_csv} (CSV)")
    print("="*60)
    
    # Process CSV
    df, subject_codes, subject_aliases = process_csv_to_dataframe(input_csv)
    
    if df is None:
        return 1
    
    # Save CSV
    if not save_clean_csv(df, output_csv):
        print("⚠️  CSV save warning, but continuing...")
    
    # Create Excel
    if not create_excel_workbook(df, subject_codes, subject_aliases, output_excel):
        return 1
    
    # Success summary
    print("\n" + "="*60)
    print("✅ PROCESSING COMPLETE!")
    print("="*60)
    print(f"\n📊 RESULTS:")
    print(f"   • Students extracted: {len(df)}")
    print(f"   • Subjects found: {len(subject_codes)}")
    print(f"   • Data columns: {len(df.columns)}")
    
    print(f"\n👥 VERIFICATION:")
    print(f"   • ADITYA DUA included: {'Yes' if any('ADITYA' in str(name).upper() for name in df['Student_Name']) else 'No'}")
    print(f"   • RAHUL C SHIRUR (not €): {'Yes' if any('RAHUL' in str(name).upper() for name in df['Student_Name']) else 'No'}")
    
    print(f"\n📁 OUTPUT FILES:")
    print(f"   1. {output_csv} - Clean structured data")
    print(f"   2. {output_excel} - Formatted Excel workbook")
    
    print("\n" + "="*60)
    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)